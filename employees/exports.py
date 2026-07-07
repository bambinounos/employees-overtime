"""
Exports XLSX: planilla de nómina mensual y reportes tabulares.

La planilla de un mes cerrado sale de los ReciboNomina (cifras congeladas).
Para el mes en curso se calcula en vivo y se marca PRELIMINAR.
"""
import io
from datetime import date
from decimal import Decimal, InvalidOperation

from django.db.models import Q

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Employee, ReciboNomina

AZUL = "1E3A5F"
GRIS_CLARO = "EEF2F7"

MESES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio',
         'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

COLUMNAS_PLANILLA = [
    ('Empleado', 28),
    ('Correo', 30),
    ('Sueldo base', 14),
    ('Horas trab.', 12),
    ('Pago horas', 14),
    ('Horas extra', 12),
    ('Pago extra', 14),
    ('Bono desempeño', 16),
    ('Comisiones', 14),
    ('Total', 14),
]


class PlanillaSinRecibosError(Exception):
    """Mes pasado sin recibos emitidos: hay que generar los recibos primero."""


def _dec(datos, clave):
    valor = datos.get(clave)
    try:
        return Decimal(str(valor)) if valor is not None else Decimal('0')
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _estilo_encabezado(ws, fila, num_cols):
    fill = PatternFill('solid', fgColor=AZUL)
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, num_cols + 1):
        celda = ws.cell(row=fila, column=col)
        celda.fill = fill
        celda.font = font
        celda.alignment = Alignment(horizontal='center')


def _filas_desde_recibos(recibos):
    for recibo in recibos:
        datos = recibo.datos
        yield [
            recibo.employee.name,
            recibo.employee.email,
            _dec(datos, 'base_salary'),
            _dec(datos, 'total_hours_worked'),
            _dec(datos, 'work_pay'),
            _dec(datos, 'total_overtime_hours'),
            _dec(datos, 'overtime_pay'),
            _dec(datos, 'performance_bonus'),
            _dec(datos, 'commission_amount'),
            recibo.total,
        ]


def _filas_en_vivo(year, month):
    primero = date(year, month, 1)
    activos = Employee.objects.filter(
        Q(end_date__isnull=True) | Q(end_date__gte=primero)).order_by('name')
    for employee in activos:
        salary = employee.calculate_salary(year, month)
        if salary is None:
            continue
        yield [
            employee.name,
            employee.email,
            salary['base_salary'],
            salary['total_hours_worked'],
            salary['work_pay'],
            salary['total_overtime_hours'],
            salary['overtime_pay'],
            salary['performance_bonus'],
            salary['commission_amount'],
            salary['total_salary'],
        ]


def generar_planilla_xlsx(year, month):
    """Devuelve (bytes_xlsx, preliminar). Lanza PlanillaSinRecibosError si el
    mes es pasado y no hay recibos emitidos."""
    hoy = date.today()
    es_mes_en_curso = (year, month) >= (hoy.year, hoy.month)

    recibos = ReciboNomina.objects.filter(year=year, month=month).select_related(
        'employee').order_by('employee__name')

    if recibos.exists():
        filas = list(_filas_desde_recibos(recibos))
        preliminar = False
    elif es_mes_en_curso:
        filas = list(_filas_en_vivo(year, month))
        preliminar = True
    else:
        raise PlanillaSinRecibosError(
            f"No hay recibos emitidos para {month:02d}/{year}. "
            "Genera los recibos del mes desde la pantalla de Nómina antes de descargar la planilla.")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Planilla {month:02d}-{year}"

    titulo = f"PLANILLA DE NÓMINA — {MESES[month]} {year}"
    if preliminar:
        titulo += " (PRELIMINAR — mes sin cerrar)"
    ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS_PLANILLA))

    fila_encabezado = 3
    for idx, (nombre, ancho) in enumerate(COLUMNAS_PLANILLA, start=1):
        ws.cell(row=fila_encabezado, column=idx, value=nombre)
        ws.column_dimensions[get_column_letter(idx)].width = ancho
    _estilo_encabezado(ws, fila_encabezado, len(COLUMNAS_PLANILLA))

    borde = Border(bottom=Side(style='thin', color='CCCCCC'))
    fila = fila_encabezado + 1
    for datos_fila in filas:
        for col, valor in enumerate(datos_fila, start=1):
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.border = borde
            if col >= 3:
                celda.number_format = '#,##0.00'
        fila += 1

    # Totales
    ws.cell(row=fila, column=1, value='TOTALES').font = Font(bold=True)
    for col in range(3, len(COLUMNAS_PLANILLA) + 1):
        letra = get_column_letter(col)
        celda = ws.cell(
            row=fila, column=col,
            value=f"=SUM({letra}{fila_encabezado + 1}:{letra}{fila - 1})" if filas else 0)
        celda.font = Font(bold=True)
        celda.number_format = '#,##0.00'
        celda.fill = PatternFill('solid', fgColor=GRIS_CLARO)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), preliminar


def rows_to_xlsx(headers, rows, sheet_name='Reporte', title=None):
    """Convierte encabezados + filas a un XLSX simple con estilo estándar."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    fila_inicio = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13, color=AZUL)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
        fila_inicio = 3

    for idx, nombre in enumerate(headers, start=1):
        ws.cell(row=fila_inicio, column=idx, value=nombre)
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(str(nombre)) + 4)
    _estilo_encabezado(ws, fila_inicio, len(headers))

    for offset, row in enumerate(rows, start=1):
        for col, valor in enumerate(row, start=1):
            ws.cell(row=fila_inicio + offset, column=col, value=valor)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
