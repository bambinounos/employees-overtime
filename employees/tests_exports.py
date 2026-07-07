"""Tests de Fase 3: planilla de nómina XLSX y exports Excel."""
import io
from datetime import date
from decimal import Decimal

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse

from openpyxl import load_workbook

from .exports import PlanillaSinRecibosError, generar_planilla_xlsx
from .models import Employee, EmployeePerformanceRecord, KPI, Salary, WorkLog
from .nomina import generar_recibo

XLSX_CT = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class PlanillaXlsxTest(TestCase):
    def setUp(self):
        self.superuser = User.objects.create_superuser('boss', 'boss@example.com', 'password')
        self.employee = Employee.objects.create(
            name='Ana Nómina', email='ana@example.com', hire_date=date(2023, 1, 1))
        Salary.objects.create(employee=self.employee, base_amount=Decimal('1600'),
                              effective_date=date(2023, 1, 1))
        for dia in range(1, 21):
            WorkLog.objects.create(employee=self.employee, date=date(2023, 1, dia),
                                   hours_worked=8)

    def test_planilla_mes_cerrado_usa_recibos(self):
        recibo = generar_recibo(self.employee, 2023, 1)
        contenido, preliminar = generar_planilla_xlsx(2023, 1)
        self.assertFalse(preliminar)

        ws = load_workbook(io.BytesIO(contenido)).active
        # Fila 4 = primera fila de datos (título en 1, encabezado en 3)
        self.assertEqual(ws.cell(row=4, column=1).value, 'Ana Nómina')
        self.assertEqual(Decimal(str(ws.cell(row=4, column=10).value)), recibo.total)

    def test_planilla_mes_pasado_sin_recibos_falla_claro(self):
        with self.assertRaises(PlanillaSinRecibosError):
            generar_planilla_xlsx(2023, 1)

    def test_planilla_mes_en_curso_es_preliminar(self):
        hoy = date.today()
        Salary.objects.create(employee=self.employee, base_amount=Decimal('1600'),
                              effective_date=hoy.replace(day=1))
        contenido, preliminar = generar_planilla_xlsx(hoy.year, hoy.month)
        self.assertTrue(preliminar)
        ws = load_workbook(io.BytesIO(contenido)).active
        self.assertIn('PRELIMINAR', ws.cell(row=1, column=1).value)

    def test_vista_planilla_solo_superuser(self):
        generar_recibo(self.employee, 2023, 1)
        url = reverse('nomina_planilla') + '?year=2023&month=1'

        user = User.objects.create_user('normal', password='password')
        Employee.objects.create(user=user, name='Normal', email='n@example.com',
                                hire_date=date(2023, 1, 1))
        self.client.login(username='normal', password='password')
        self.assertEqual(self.client.get(url).status_code, 403)

        self.client.login(username='boss', password='password')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], XLSX_CT)

    def test_vista_planilla_sin_recibos_redirige_con_error(self):
        self.client.login(username='boss', password='password')
        response = self.client.get(reverse('nomina_planilla') + '?year=2022&month=6')
        self.assertEqual(response.status_code, 302)


class PerformanceReportXlsxTest(TestCase):
    def setUp(self):
        self.superuser = User.objects.create_superuser('boss', 'boss@example.com', 'password')
        self.employee = Employee.objects.create(
            name='Kpi Man', email='kpi@example.com', hire_date=date(2023, 1, 1))
        kpi = KPI.objects.create(name='Puntualidad', measurement_type='percentage',
                                 target_value=Decimal('90'))
        EmployeePerformanceRecord.objects.create(
            employee=self.employee, kpi=kpi, date=date(2023, 1, 31),
            actual_value=Decimal('95'), target_met=True, bonus_awarded=Decimal('50'))

    def test_export_xlsx(self):
        self.client.login(username='boss', password='password')
        response = self.client.get(reverse('performance_report'),
                                   {'employee_id': self.employee.id, 'format': 'xlsx'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], XLSX_CT)
        ws = load_workbook(io.BytesIO(response.content)).active
        self.assertEqual(ws.cell(row=3, column=2).value, 'KPI')
        self.assertEqual(ws.cell(row=4, column=2).value, 'Puntualidad')

    def test_export_csv_sigue_funcionando(self):
        self.client.login(username='boss', password='password')
        response = self.client.get(reverse('performance_report'),
                                   {'employee_id': self.employee.id, 'format': 'csv'})
        self.assertEqual(response.status_code, 200)
        self.assertIn('Puntualidad', response.content.decode())
